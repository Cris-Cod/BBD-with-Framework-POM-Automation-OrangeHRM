import openpyxl
import os


def getTestData():
    path = os.getcwd()
    book = openpyxl.load_workbook(f"{path}\\data.xlsx")
    sheet = book.active
    data_comb = []
    cell = sheet.cell(row=2, column=1)

    for i in range(2, sheet.max_row+1):
        row = []
        for j in range(1, sheet.max_column+1):
            row.append(sheet.cell(row=i, column=j).value)
        data_comb.append(row)

    return data_comb

print(getTestData())

